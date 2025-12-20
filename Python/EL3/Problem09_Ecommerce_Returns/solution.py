import openpyxl
import csv

# Read returns data
wb = openpyxl.load_workbook('returns.xlsx')
ws = wb.active

# Valid refund modes
valid_modes = ['UPI', 'CARD', 'WALLET']

# Lists to store valid and invalid returns
valid_returns = []
invalid_returns = []

# Process each return (skip header)
for row in ws.iter_rows(min_row=2, values_only=True):
    order_id = row[0]
    customer = row[1]
    mode = row[2]
    amount = row[3]
    
    error = None
    
    # Validate refund mode
    if mode not in valid_modes:
        error = 'Invalid RefundMode'
    # Validate amount
    elif amount <= 0:
        error = 'Invalid Amount'
    
    if error:
        invalid_returns.append([order_id, customer, mode, amount, error])
    else:
        valid_returns.append([order_id, customer, mode, amount])

# Write valid returns to CSV
with open('returns_clean.csv', 'w', newline='') as outfile:
    writer = csv.writer(outfile)
    writer.writerow(['OrderID', 'CustomerName', 'RefundMode', 'Amount'])
    writer.writerows(valid_returns)

# Write invalid returns to Excel with error reasons
wb_err = openpyxl.Workbook()
ws_err = wb_err.active
ws_err.append(['OrderID', 'CustomerName', 'RefundMode', 'Amount', 'ErrorReason'])

for row in invalid_returns:
    ws_err.append(row)

wb_err.save('returns_error_log.xlsx')

print("Returns validation completed successfully!")
