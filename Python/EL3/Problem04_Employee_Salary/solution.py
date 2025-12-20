import openpyxl

# Read employee data
wb = openpyxl.load_workbook('emp_data.xlsx')
ws = wb.active

# Create output workbook
wb_out = openpyxl.Workbook()
ws_out = wb_out.active

# Write headers
ws_out.append(['EmpID', 'Name', 'GrossSalary'])

# Process each employee (skip header)
for row in ws.iter_rows(min_row=2, values_only=True):
    emp_id = row[0]
    name = row[1]
    basic = row[3]
    
    # Calculate HRA (10%) and DA (18%)
    hra = basic * 0.10
    da = basic * 0.18
    
    # Calculate gross salary
    gross = basic + hra + da
    
    ws_out.append([emp_id, name, round(gross, 2)])

# Save output
wb_out.save('emp_salary.xlsx')
print("Employee salary file created successfully!")
