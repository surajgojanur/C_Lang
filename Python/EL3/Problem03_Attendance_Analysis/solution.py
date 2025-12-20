import openpyxl

# Read attendance.xlsx
wb = openpyxl.load_workbook('attendance.xlsx')
ws = wb.active

# Create new workbook for report
wb_out = openpyxl.Workbook()
ws_out = wb_out.active

# Write headers
ws_out.append(['RollNo', 'Name', 'ClassesHeld', 'ClassesAttended', 'AttendancePercentage', 'AttendanceStatus'])

# Process each student (skip header row)
for row in ws.iter_rows(min_row=2, values_only=True):
    roll_no = row[0]
    name = row[1]
    held = row[2]
    attended = row[3]
    
    # Calculate attendance percentage
    percentage = (attended / held) * 100
    
    # Determine status
    if percentage < 75:
        status = 'Shortage'
    else:
        status = 'OK'
    
    ws_out.append([roll_no, name, held, attended, round(percentage, 2), status])

# Save report
wb_out.save('attendance_report.xlsx')
print("Attendance report created successfully!")
