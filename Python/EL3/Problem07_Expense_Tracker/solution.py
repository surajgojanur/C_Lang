import openpyxl

# Read expenses data
wb = openpyxl.load_workbook('expenses.xlsx')
ws = wb.active

# Dictionary to store category-wise totals
category_totals = {}

# Process expenses (skip header)
for row in ws.iter_rows(min_row=2, values_only=True):
    category = row[1]
    amount = row[2]
    
    # Add to category total
    if category in category_totals:
        category_totals[category] += amount
    else:
        category_totals[category] = amount

# Create summary workbook
wb_out = openpyxl.Workbook()
ws_out = wb_out.active

# Write headers
ws_out.append(['Category', 'MonthlyTotal'])

# Write category totals
for category, total in category_totals.items():
    ws_out.append([category, total])

# Save summary
wb_out.save('monthly_summary.xlsx')
print("Monthly summary created successfully!")
